import serial

import time

import openpyxl

from openpyxl.styles import Font, Alignment

 

def lire_donnees_serial(port, baudrate, timeout=5):

    """

    Lit les données du port série avec gestion des erreurs.

    """

    print(f"Connexion au port série {port} avec un baudrate de {baudrate}...")

    try:

        with serial.Serial(port, baudrate, timeout=timeout) as ser:

            print("Connecté au port série.\n")

            lignes = []

            while True:

                raw_data = ser.readline()  # Lire une ligne brute

                if raw_data:

                    print(f"Données brutes reçues : {raw_data}")

                    try:

                        # Tentative de décodage des données

                        ligne = raw_data.decode('utf-8', errors='ignore').strip()

                        print(f"Données décodées : {ligne}")

                        lignes.append(ligne)

 

                        # Si une ligne contient "END", fin de réception (ou adaptez selon votre protocole)

                        if "END" in ligne:

                            print("Fin des données détectée (mot-clé 'END').")

                            break

                    except Exception as decode_error:

                        print(f"Erreur lors du décodage : {decode_error}")

                else:

                    print("Aucune donnée reçue (timeout atteint).")

                time.sleep(0.5)  # Pause pour éviter une boucle trop rapide

    except serial.SerialException as e:

        print(f"Erreur de connexion au port série : {e}")

        return []

    return lignes

 

def extraire_resultats(donnees):

    """

    Extrait les résultats des tests des données reçues.

    """

    resultats = []

    for ligne in donnees:

        if "Test" in ligne and ("Pass" in ligne or "Fail" in ligne):

            # Exemple de format attendu : "Test 1: Pass" ou "Test 2: Fail"

            test_id, statut = ligne.split(":", 1)  # Permet d'éviter une erreur si plusieurs ':' existent

            resultats.append({

                "Test": test_id.strip(),

                "Statut": statut.strip()

            })

    return resultats

 

def creer_tableur_excel(resultats, fichier="resultats_tests.xlsx"):

    """

    Crée un fichier Excel avec les résultats des tests.

    """

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Résultats des Tests"

 

    # En-têtes

    ws.append(["Test", "Statut"])

    ws["A1"].font = ws["B1"].font = Font(bold=True)

    ws["A1"].alignment = ws["B1"].alignment = Alignment(horizontal="center")

 

    # Ajout des résultats

    for res in resultats:

        ws.append([res["Test"], res["Statut"]])

 

    # Ajustement de la largeur des colonnes

    for col in ws.columns:

        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)

        col_letter = col[0].column_letter  # Récupère la lettre de la colonne

        ws.column_dimensions[col_letter].width = max_length + 2

 

    # Sauvegarde du fichier

    wb.save(fichier)

    print(f"Fichier Excel généré : {fichier}")

 

if __name__ == "__main__":

    # Configuration du port série

    PORT = "COM5"  # Changez selon votre système

    BAUDRATE = 115200

    TIMEOUT = 5

 

    # Lecture des données

    print("Démarrage de la lecture des données série...")

    donnees = lire_donnees_serial(PORT, BAUDRATE, timeout=TIMEOUT)

 

    if donnees:

        # Extraction des résultats

        print("\nExtraction des résultats des données reçues...")

        resultats = extraire_resultats(donnees)

 

        # Création du fichier Excel

        if resultats:

            creer_tableur_excel(resultats)

        else:

            print("Aucun résultat à enregistrer dans le fichier Excel.")

    else:

        print("Aucune donnée série reçue ou traitement interrompu.")